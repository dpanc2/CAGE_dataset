#!/usr/bin/env python3
"""Join the GEO/ENA manifest to the published 31-heart subject mapping."""

import argparse
import csv
import re
from pathlib import Path

from openpyxl import load_workbook


def norm(value):
    return "" if value is None else str(value).strip()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--manifest", type=Path, required=True)
    parser.add_argument("--supplement", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    with args.manifest.open(newline="") as handle:
        manifest = list(csv.DictReader(handle, delimiter="\t"))

    workbook = load_workbook(args.supplement, read_only=True, data_only=True)
    sheet = workbook["Table 1 - Sample statistics"]
    rows = sheet.iter_rows(values_only=True)
    header = [norm(x) for x in next(rows)]
    published = [dict(zip(header, row)) for row in rows if any(x is not None for x in row)]

    if len(manifest) != 109 or len(published) != 109:
        raise SystemExit(
            f"Expected 109 GEO and 109 published rows; got {len(manifest)} and {len(published)}"
        )

    chamber_names = {
        "LA": "left_atrium",
        "LV": "left_ventricle",
        "RA": "right_atrium",
        "RV": "right_ventricle",
        "RA-sSAN": "sSAN",
        "RA-iSAN": "iSAN",
    }
    output = []
    for position, (geo, pub) in enumerate(zip(manifest, published), start=1):
        match = re.search(r"Sample\s+(\d+)", geo["sample_title"])
        if not match or int(match.group(1)) != position:
            raise SystemExit(f"Unexpected GEO sample order at row {position}: {geo['sample_title']}")

        expected_chamber = chamber_names.get(norm(pub["heart"]), norm(pub["heart"]))
        observed_chamber = geo["chamber"].replace(" ", "_")
        checks = {
            "sex": geo["gender"] == norm(pub["sex"]),
            "age": geo["age"] == norm(pub["age"]),
            "state": (
                geo["state"].lower() == norm(pub["group"]).lower()
                or (
                    geo["state"].lower().startswith("failing_")
                    and norm(pub["group"]).lower() == "failing"
                )
            ),
            "chamber": observed_chamber == expected_chamber,
        }
        failed = [key for key, passed in checks.items() if not passed]
        if failed:
            raise SystemExit(f"Metadata mismatch at Sample {position}: {', '.join(failed)}")

        subject = int(pub["subject #"])
        output.append({
            "donor_id": f"subject_{subject:02d}",
            "subject_number": subject,
            "run_accession": geo["run_accession"],
            "geo_accession": geo["geo_accession"],
            "sample_id": norm(pub["id"]),
            "chamber": norm(pub["heart"]),
            "sex": norm(pub["sex"]),
            "age": norm(pub["age"]),
            "state": norm(pub["group"]),
            "cardiomyopathy_type": norm(pub["cardiomyopathy type"]),
            "fastq_name": f"{geo['run_accession']}.fastq.gz",
            "fastq_url": geo["fastq_url"],
            "fastq_md5": geo["fastq_md5"],
            "fastq_bytes": geo["fastq_bytes"],
        })

    donors = {row["donor_id"] for row in output}
    if len(donors) != 31:
        raise SystemExit(f"Expected 31 donors; got {len(donors)}")

    args.output.parent.mkdir(parents=True, exist_ok=True)
    fields = list(output[0])
    with args.output.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, delimiter="\t")
        writer.writeheader()
        writer.writerows(output)
    print(f"Wrote {len(output)} samples from {len(donors)} donors to {args.output}")


if __name__ == "__main__":
    main()
