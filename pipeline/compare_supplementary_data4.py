#!/usr/bin/env python3
"""Compare pilot allele counts with a MIXALIME Supplementary Data 4 sheet."""

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


def read_counts(paths):
    combined = {}
    runs = []
    for path in paths:
        run = path.name.split(".allele_counts.tsv", 1)[0]
        runs.append(run)
        with path.open(newline="") as handle:
            for row in csv.DictReader(handle, delimiter="\t"):
                key = (row["chrom"], int(row["pos"]), row["ref"], row["alt"])
                item = combined.setdefault(key, {"runs": {}})
                item["runs"][run] = {
                    "ref": int(row["ref_count"]),
                    "alt": int(row["alt_count"]),
                    "other": int(row["other_count"]),
                    "depth": int(row["depth"]),
                }
    return combined, runs


def read_supplement(path, sheet_name):
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(
            f"Unknown sheet {sheet_name!r}; choose from: {', '.join(workbook.sheetnames)}"
        )
    rows = workbook[sheet_name].iter_rows(values_only=True)
    header = next(rows)
    columns = {name: i for i, name in enumerate(header) if name is not None}
    required = {"#chr", "end", "id", "ref", "alt", "n_reps", "pref_allele", "comb_es", "fdr_comb_pval"}
    missing = required - columns.keys()
    if missing:
        raise SystemExit(f"Sheet {sheet_name!r} lacks columns: {', '.join(sorted(missing))}")

    result = {}
    for row in rows:
        if row[columns["#chr"]] is None:
            continue
        key = (
            str(row[columns["#chr"]]),
            int(row[columns["end"]]),  # Supplement is BED-like; end equals VCF POS.
            str(row[columns["ref"]]),
            str(row[columns["alt"]]),
        )
        result[key] = {
            "id": row[columns["id"]] or ".",
            "n_reps": row[columns["n_reps"]],
            "pref": row[columns["pref_allele"]],
            "effect": row[columns["comb_es"]],
            "fdr": row[columns["fdr_comb_pval"]],
        }
    return result


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--supplement", type=Path, required=True)
    parser.add_argument("--sheet", default="BetaNB")
    parser.add_argument("--counts", type=Path, nargs="+", required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    counts, runs = read_counts(args.counts)
    published = read_supplement(args.supplement, args.sheet)
    fields = ["chrom", "pos", "ref", "alt"]
    for run in runs:
        fields += [f"{run}_ref", f"{run}_alt", f"{run}_depth"]
    fields += [
        "total_ref", "total_alt", "total_depth", "alt_fraction", "pilot_pref_allele",
        "overlap", "allele_orientation", "rsid", "published_n_reps",
        "published_pref_allele", "published_effect", "published_fdr",
        "published_significant_5pct", "direction_concordant",
    ]

    args.output.parent.mkdir(parents=True, exist_ok=True)
    overlap_n = significant_n = concordant_n = 0
    with args.output.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, delimiter="\t")
        writer.writeheader()
        for key in sorted(counts, key=lambda x: (x[0], x[1], x[2], x[3])):
            chrom, pos, ref, alt = key
            item = counts[key]
            total_ref = sum(item["runs"].get(r, {}).get("ref", 0) for r in runs)
            total_alt = sum(item["runs"].get(r, {}).get("alt", 0) for r in runs)
            total_depth = total_ref + total_alt
            pilot_pref = "ref" if total_ref > total_alt else "alt" if total_alt > total_ref else "tie"
            pub = published.get(key)
            orientation = "direct"
            if pub is None:
                pub = published.get((chrom, pos, alt, ref))
                orientation = "swapped" if pub is not None else "."

            row = {"chrom": chrom, "pos": pos, "ref": ref, "alt": alt}
            for run in runs:
                values = item["runs"].get(run, {})
                row.update({
                    f"{run}_ref": values.get("ref", 0),
                    f"{run}_alt": values.get("alt", 0),
                    f"{run}_depth": values.get("depth", 0),
                })
            row.update({
                "total_ref": total_ref,
                "total_alt": total_alt,
                "total_depth": total_depth,
                "alt_fraction": f"{total_alt / total_depth:.6f}" if total_depth else ".",
                "pilot_pref_allele": pilot_pref,
                "overlap": "yes" if pub else "no",
                "allele_orientation": orientation,
                "rsid": pub["id"] if pub else ".",
                "published_n_reps": pub["n_reps"] if pub else ".",
                "published_pref_allele": pub["pref"] if pub else ".",
                "published_effect": pub["effect"] if pub else ".",
                "published_fdr": pub["fdr"] if pub else ".",
                "published_significant_5pct": "yes" if pub and float(pub["fdr"]) < 0.05 else "no",
                "direction_concordant": ".",
            })
            if pub:
                overlap_n += 1
                pub_pref = pub["pref"]
                if orientation == "swapped":
                    pub_pref = "alt" if pub_pref == "ref" else "ref"
                if float(pub["fdr"]) < 0.05:
                    significant_n += 1
                if pilot_pref != "tie":
                    row["direction_concordant"] = "yes" if pilot_pref == pub_pref else "no"
                    concordant_n += row["direction_concordant"] == "yes"
            writer.writerow(row)

    print(f"Pilot loci: {len(counts)}")
    print(f"Overlap with {args.sheet}: {overlap_n}")
    print(f"Published FDR < 0.05 among overlaps: {significant_n}")
    print(f"Direction-concordant overlaps: {concordant_n}")
    print(f"Output: {args.output}")


if __name__ == "__main__":
    main()
